import csv
import itertools

from django.core.management.base import BaseCommand, no_translations, CommandError
from django.db import transaction
from openpyxl import load_workbook

from mainapp.models import (Atividade, DatasFuncionamento, Turma,
                            TurmaAtividades, Aluno, Periodo, NotaAluno, TurmaAlunos)


periodo = Periodo.objects.get(is_atual=1)


class Command(BaseCommand):
    help = 'Exporta um arquivo CSV a partir de uma planilha'

    requires_migrations_checks = True
    requires_system_checks = True
    output_transaction = True

    def add_arguments(self, parser):
        parser.add_argument('excel_file',
                            type=str, help='Planilha Excel')

    @no_translations
    @transaction.atomic()
    def handle(self, *args, **options):
        wb = load_workbook(
            filename=options['excel_file'], data_only=True)
        sheets = wb.sheetnames
        for sheetname in sheets:
            sheet = wb[sheetname]

            # Checando se as âncoras do modelo padrão estão presentes
            if sheet['B4'].value is None or sheet['A9'].value is None:
                raise CommandError('Arquivo não conforma com modelo padrão!\n Âncoras: %s, %s' % (
                    sheet['B4'].value, sheet['A9'].value))
            linha_notas = 4
            while sheet['A%d' % linha_notas].value is None or sheet['A%d' % linha_notas].value.strip() != 'Nome dos Alunos' or linha_notas > 15:
                linha_notas += 1
            if sheet['A%d' % linha_notas].value.strip() != 'Nome dos Alunos' or sheet['A%d' % (linha_notas+1)].value is not None or sheet['A%d' % (linha_notas+2)].value is None:
                raise CommandError(
                    'Arquivo não conforma com modelo padrão!\n Âncora: %s' % sheet['A%d' % linha_notas].value)
            self.stdout.write(self.style.SUCCESS(
                '\t[0/7] Documento bem-formatado'))

            nome_turma = ''.join(sheet['B4'].value.strip().split(' '))
            try:
                turma = Turma.objects.get(
                    nome_turma=nome_turma, id_periodo=periodo)
                self.stdout.write(self.style.SUCCESS(
                    '\t[1/7] Turma encontrada %s:' % turma.nome_turma))
            except Turma.DoesNotExist:
                raise CommandError(
                    'Turma %s não encontrada no banco de dados!' % nome_turma)

            alunos = list()
            for num in itertools.count(linha_notas+2):
                if sheet['A%d' % num].value is None:
                    break
                value = sheet['A%d' % num].value.strip().upper()
                if value == '':
                    break
                try:
                    aluno = TurmaAlunos.objects.get(
                        id_aluno__nome_aluno=value, id_turma=turma)
                    alunos.append({'ent': aluno, 'pos': num})
                except TurmaAlunos.DoesNotExist:
                    raise CommandError(
                        'Aluno %s não existe no banco de dados!' % value)
                except TurmaAlunos.MultipleObjectsReturned:
                    raise CommandError(
                        'Aluno %s está ambíguo no banco de dados!' % value)
            self.stdout.write(self.style.SUCCESS(
                '\t[2/6] Identificados %d alunos' % len(alunos)))

            coluna_notas = 'A'
            celula = sheet['%s%d' % (coluna_notas, linha_notas)].value
            # O coluna_notas<20 é apenas um sanity check para evitar timeouts
            while (sheet['%s%d' % (coluna_notas, linha_notas)].value is None or sheet['%s%d' % (coluna_notas, linha_notas)].value.strip() != 'Total (100 pts)') and (ord(coluna_notas) - ord('A')) < 20:
                coluna_notas = chr(ord(coluna_notas) + 1)
                celula = sheet['%s%d' % (coluna_notas, linha_notas)].value
            if celula.strip() != 'Total (100 pts)':
                raise CommandError(
                    'Coluna com notas finais não foi encontrada!')
            self.stdout.write(self.style.SUCCESS(
                '\t[3/6] Encontrada a coluna com as notas finais'))

            atividade_final = Atividade(data_funcionamento=DatasFuncionamento.objects.get(
                data_funcionamento__year=2019, data_funcionamento__month=7, data_funcionamento__day=20), nome='Nota Final', descricao='Lançamento automático', valor_total=100)
            atividade_final.save()
            turmaatividade = TurmaAtividades(
                id_turma=turma, id_atividade=atividade_final)
            turmaatividade.save()

            for aluno in alunos:
                valor_nota = round(
                    float(sheet['%s%d' % (coluna_notas, aluno['pos'])].value), 1)
                nota = NotaAluno(
                    id_turma_aluno=aluno['ent'], id_atividades_turma=turmaatividade, valor_nota=valor_nota)
                nota.save()
                self.stdout.write('\t\tNota: %s (%d), Aluno: "%s" (%d), Atividade: "%s" (%d)' % (nota.valor_nota, nota.id_nota, nota.id_turma_aluno.id_aluno.nome_aluno,
                                                                                                 nota.id_turma_aluno.id_aluno.id_aluno, nota.id_atividades_turma.id_atividade.nome, nota.id_atividades_turma.id_atividade.id_atividade))
            self.stdout.write(self.style.SUCCESS(
                '\t[4/6] Lançadas %d notas' % len(alunos)))

            outras_notas = NotaAluno.objects.filter(id_turma_aluno__id_turma=turma).exclude(
                id_atividades_turma__id_atividade=atividade_final)
            self.stdout.write(self.style.WARNING(
                '\t[5/6] Deletando %d notas antigas' % outras_notas.count()))
            outras_notas.delete()

            outras_atividades = TurmaAtividades.objects.filter(
                id_turma=turma).exclude(id_atividade=atividade_final)
            self.stdout.write(self.style.WARNING(
                '\t[6/6] Deletando %d atividades antigas' % outras_atividades.count()))
            for old_turmaatividade in outras_atividades:
                old_ativ = old_turmaatividade.id_atividade
                old_turmaatividade.delete()
                old_ativ.delete()

    def char_range(self, c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)


# Funções-gambiarra para encontrar e corrigir erros nas planilhas

def acha(aluno, turma):
    turmaalunos = TurmaAlunos.objects.filter(
        id_aluno__nome_aluno__icontains=aluno, id_turma__nome_turma=turma, id_turma__id_periodo__is_atual=1)
    return list(map(lambda ta: (ta.id_aluno.nome_aluno, ta.id_turma.nome_turma), turmaalunos))


def turmas(aluno):
    turmaalunos = TurmaAlunos.objects.filter(
        id_aluno__nome_aluno__icontains=aluno, id_turma__id_periodo__is_atual=1)
    return list(map(lambda ta: (ta.id_aluno.nome_aluno, ta.id_turma.nome_turma), turmaalunos))


def mudaturma(oldturma, newturma, nome):
    aluno = Aluno.objects.get(nome_aluno__icontains=nome)
    velho = TurmaAlunos.objects.get(
        id_aluno=aluno, id_turma__nome_turma=oldturma, id_turma__id_periodo__is_atual=1)
    novo = TurmaAlunos(id_aluno=aluno, id_turma=Turma.objects.get(nome_turma=newturma, id_periodo__is_atual=1),
                       id_pagamento=velho.id_pagamento, aprovado=velho.aprovado, liberacao=velho.liberacao)
    velho.delete()
    novo.save()
    return turmas(nome)


def limpa():
    NotaAluno.objects.filter(
        id_turma_aluno__id_turma__id_periodo__is_atual=1).delete()
