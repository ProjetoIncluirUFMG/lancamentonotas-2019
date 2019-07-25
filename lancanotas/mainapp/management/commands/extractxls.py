import csv
import itertools

from django.core.management.base import BaseCommand, no_translations, CommandError
from django.db import transaction
from openpyxl import load_workbook
from unidecode import unidecode

from mainapp.models import (Atividade, DatasFuncionamento, Turma,
                            TurmaAtividades, Aluno, Periodo, NotaAluno, TurmaAlunos)


periodo = Periodo.objects.get(is_atual=1)


class Command(BaseCommand):
    help = 'Exporta um arquivo CSV a partir de uma planilha'

    requires_migrations_checks = True
    requires_system_checks = True
    output_transaction = True

    def add_arguments(self, parser):
        parser.add_argument('excel_file',
                            type=str, help='Planilha Excel')

    @no_translations
    @transaction.atomic()
    def handle(self, *args, **options):
        wb = load_workbook(
            filename=options['excel_file'], data_only=True)
        sheets = wb.sheetnames
        for sheetname in sheets:
            if not Turma.objects.filter(nome_turma=sheetname, id_periodo=periodo).exists():
                raise CommandError(
                    'Turma %s não encontrada no banco de dados!' % sheetname)
            turma = Turma.objects.get(
                nome_turma=sheetname, id_periodo=periodo)
            self.stdout.write(self.style.SUCCESS(
                'Turma encontrada %s:' % turma.nome_turma))

            # Checando se as âncoras do modelo padrão estão presentes
            sheet = wb[sheetname]
            if not (sheet['B6'].value is not None and sheet['B6'].value.strip() == 'Nome da Atividade' and sheet['B8'].value is not None and sheet['B8'].value.strip() == 'Nome do Aluno'):
                raise CommandError('Arquivo não conforma com modelo padrão!\n Âncoras: %s, %s' % (
                    sheet['B6'].value, sheet['B8'].value))
            self.stdout.write(self.style.SUCCESS(
                '\t[1/7] Documento bem-formatado'))

            alunos = list()
            for num in itertools.count(10):
                if sheet['B%d' % num].value is None:
                    break
                value = unidecode(sheet['B%d' % num].value.strip().upper())
                if value == '':
                    break
                try:
                    aluno = TurmaAlunos.objects.get(
                        id_aluno__nome_aluno=value, id_turma=turma)
                    alunos.append({'ent': aluno, 'pos': num})
                except TurmaAlunos.DoesNotExist:
                    raise CommandError(
                        'Aluno %s não existe no banco de dados!' % value)
            self.stdout.write(self.style.SUCCESS(
                '\t[2/6] Identificados %d alunos' % len(alunos)))

            coluna_notas = 'C'
            celula = sheet['%s6' % coluna_notas].value
            # O coluna_notas<20 é apenas um sanity check para evitar timeouts
            while (sheet['%s6' % coluna_notas].value is None or sheet['%s6' % coluna_notas].value.strip() != 'Total') and (ord(coluna_notas) - ord('C')) < 20:
                coluna_notas = chr(ord(coluna_notas) + 1)
                celula = sheet['%s6' % coluna_notas].value
            if celula.strip() != 'Total':
                raise CommandError(
                    'Coluna com notas finais não foi encontrada!')
            self.stdout.write(self.style.SUCCESS(
                '\t[3/6] Encontrada a coluna com as notas finais'))

            atividade_final = Atividade(data_funcionamento=DatasFuncionamento.objects.get(
                data_funcionamento__year=2019, data_funcionamento__month=7, data_funcionamento__day=20), nome='Nota Final', descricao='Lançamento automático', valor_total=100)
            atividade_final.save()
            turmaatividade = TurmaAtividades(
                id_turma=turma, id_atividade=atividade_final)
            turmaatividade.save()

            for aluno in alunos:
                valor_nota = round(
                    float(sheet['%s%d' % (coluna_notas, aluno['pos'])].value), 1)
                nota = NotaAluno(
                    id_turma_aluno=aluno['ent'], id_atividades_turma=turmaatividade, valor_nota=valor_nota)
                nota.save()
                self.stdout.write('\t\tNota: %s (%d), Aluno: "%s" (%d), Atividade: "%s" (%d)' % (nota.valor_nota, nota.id_nota, nota.id_turma_aluno.id_aluno.nome_aluno,
                                                                                                 nota.id_turma_aluno.id_aluno.id_aluno, nota.id_atividades_turma.id_atividade.nome, nota.id_atividades_turma.id_atividade.id_atividade))
            self.stdout.write(self.style.SUCCESS(
                '\t[4/6] Lançadas %d notas' % len(alunos)))

            outras_notas = NotaAluno.objects.filter(id_turma_aluno__id_turma=turma).exclude(
                id_atividades_turma__id_atividade=atividade_final)
            self.stdout.write(self.style.WARNING(
                '\t[5/6] Deletando %d notas antigas' % outras_notas.count()))
            outras_notas.delete()

            outras_atividades = TurmaAtividades.objects.filter(
                id_turma=turma).exclude(id_atividade=atividade_final)
            self.stdout.write(self.style.WARNING(
                '\t[6/6] Deletando %d atividades antigas' % outras_atividades.count()))
            for old_turmaatividade in outras_atividades:
                old_ativ = old_turmaatividade.id_atividade
                old_turmaatividade.delete()
                old_ativ.delete()

    def char_range(self, c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)
